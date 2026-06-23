from __future__ import annotations

import asyncio
import csv
import logging
from pathlib import Path
from typing import Any

import openpyxl

from app.schemas.file_analysis import SheetData

logger = logging.getLogger(__name__)

_SUPPORTED = {".xlsx", ".xls", ".csv"}
_MAX_ROWS = 500


def _read_xlsx_sync(file_path: str) -> list[SheetData]:
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    sheets: list[SheetData] = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        all_rows = list(ws.iter_rows(values_only=True))
        if not all_rows:
            continue
        headers = [str(h) if h is not None else "" for h in all_rows[0]]
        data_rows: list[dict[str, Any]] = []
        for row in all_rows[1 : _MAX_ROWS + 1]:
            data_rows.append(
                {headers[i]: (str(v) if v is not None else "") for i, v in enumerate(row)}
            )
        sheets.append({"sheet": sheet_name, "headers": headers, "rows": data_rows})
    wb.close()
    return sheets


def _read_csv_sync(file_path: str) -> list[SheetData]:
    with open(file_path, newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        headers = list(reader.fieldnames or [])
        rows = [dict(row) for _, row in zip(range(_MAX_ROWS), reader)]
    return [{"sheet": "Sheet1", "headers": headers, "rows": rows}]


async def parse_excel(file_path: str) -> list[SheetData]:
    """Extract sheet data from an Excel (.xlsx/.xls) or CSV file.

    Returns a list of dicts with 'sheet', 'headers', and 'rows' keys.
    Raises FileNotFoundError for missing files, ValueError for unsupported types.
    """
    path = Path(file_path)
    if not path.exists():
        raise FileNotFoundError(f"File not found: {file_path}")

    suffix = path.suffix.lower()
    if suffix not in _SUPPORTED:
        raise ValueError(f"Unsupported file type: {path.suffix!r}. Expected one of {_SUPPORTED}.")

    loop = asyncio.get_event_loop()
    if suffix == ".csv":
        sheets = await loop.run_in_executor(None, _read_csv_sync, str(path))
    else:
        sheets = await loop.run_in_executor(None, _read_xlsx_sync, str(path))

    logger.debug("parse_excel: extracted %d sheet(s) from %r", len(sheets), path.name)
    return sheets
